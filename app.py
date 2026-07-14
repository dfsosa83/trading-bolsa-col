"""
BVC Fixed Income Dashboard
Data is fetched live from the BVC API on every session (cached 1 h).
No local files required — works on Streamlit Community Cloud.
"""

import io
import sys
from datetime import date
from pathlib import Path

import altair as alt
import openpyxl
import pandas as pd
import requests
import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent))
from download_bvc_daily import fetch_report_list, download_xlsx_bytes
from reports.parsers import rf_mercado_secundario, rf_por_tipo_inver
from reports.generate_report import generate_to_bytes

# ── Config ─────────────────────────────────────────────────────────────────────

_EXT_DEBT_INSTR = "PUBLIC EXTERNAL DEBT BONDS DOLLAR DENOMINATED"
_HISTORY_DIR    = Path(__file__).resolve().parent / "data" / "history"

st.set_page_config(
    page_title="BVC · Bonos Ext. USD",
    page_icon="📊",
    layout="wide",
)

# ── Cached data functions ──────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def _get_reports() -> list[dict]:
    """Fetch available daily reports from BVC API. Re-checked every hour."""
    return fetch_report_list()


@st.cache_data(ttl=3600, show_spinner=False)
def _get_trm() -> dict:
    """
    Fetch the current TRM (Tasa Representativa del Mercado) from the
    Superintendencia Financiera via datos.gov.co (no API key required).
    Returns dict with 'valor' (float COP/USD) and 'vigencia' (str date).
    Falls back to a safe default if the request fails.
    """
    try:
        resp = requests.get(
            "https://www.datos.gov.co/resource/32sa-8pi3.json"
            "?$limit=1&$order=vigenciadesde+DESC",
            timeout=5,
        )
        resp.raise_for_status()
        row = resp.json()[0]
        return {
            "valor":    float(row["valor"]),
            "vigencia": row["vigenciadesde"][:10],
        }
    except Exception:
        return {"valor": 4_000.0, "vigencia": "N/D"}


def _parse(xlsx_bytes: bytes, report_date: str):
    """Parse xlsx bytes into all dashboard data. Shared by both cached loaders."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ms_rows  = list(wb["RF-Mercado Secundario"].iter_rows(values_only=True))
    pti_rows = list(wb["RF-PorTipoInver"].iter_rows(values_only=True))
    wb.close()

    ms_data  = rf_mercado_secundario.parse(ms_rows)
    pti_data = rf_por_tipo_inver.parse(pti_rows)

    # ── BGLT bonds ────────────────────────────────────────────────────────────
    df_reg = ms_data["pub_registro"].copy()
    mask   = df_reg["descripcion"].str.lower().str.contains("ext dolares", na=False)
    df_bonds = df_reg[mask].copy()
    df_bonds["total_monto"] = df_bonds["cv_monto"].fillna(0) + df_bonds["sim_monto"].fillna(0)

    bond_display = df_bonds[[
        "nemotecnico", "cv_monto", "fec_vcto",
        "cv_num_opes",
        "cv_tasa_min", "cv_tasa_max", "cv_tasa_cierre",
    ]].rename(columns={
        "nemotecnico":    "Nemotécnico",
        "cv_monto":       "CV Monto",
        "fec_vcto":       "Vencimiento",
        "cv_num_opes":    "# Oper.",
        "cv_tasa_min":    "Tasa Mín",
        "cv_tasa_max":    "Tasa Máx",
        "cv_tasa_cierre": "Tasa Cierre",
    }).reset_index(drop=True)
    bond_display["CV Monto"] = bond_display["CV Monto"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")

    # ── Buyers / Sellers ──────────────────────────────────────────────────────
    def _find_ext_debt_col(df: pd.DataFrame) -> str:
        if _EXT_DEBT_INSTR in df.columns:
            return _EXT_DEBT_INSTR
        for col in df.columns:
            if "external" in col.lower() and "dollar" in col.lower():
                return col
        for col in df.columns:
            if col not in ("sector", "total"):
                return col
        raise KeyError(f"Cannot find external-debt column in {list(df.columns)}")

    def _side(key, col_label):
        df  = pti_data[key]
        col = _find_ext_debt_col(df)
        out = (
            df[["sector", col]]
            .rename(columns={"sector": "Sector", col: col_label})
            .sort_values(col_label, ascending=False)
            .reset_index(drop=True)
        )
        out[col_label] = out[col_label].round().astype("Int64")
        return out

    buyers  = _side("buyers",  "Monto Comprado")
    sellers = _side("sellers", "Monto Vendido")
    total   = df_bonds["total_monto"].sum()

    report_bytes = generate_to_bytes(xlsx_bytes, report_date).getvalue()
    return bond_display, buyers, sellers, total, report_bytes


@st.cache_data(show_spinner=False)
def _load_from_file(xlsx_path: str, report_date: str):
    """Load dashboard data from a saved history file (no network call)."""
    return _parse(Path(xlsx_path).read_bytes(), report_date)


@st.cache_data(show_spinner=False)
def _load(url: str, report_date: str):
    """Download xlsx for the given report URL, parse it, and build all dashboard data."""
    return _parse(download_xlsx_bytes(url), report_date)


@st.cache_data(show_spinner=False)
def _build_yield_history_df(_files_key: tuple) -> pd.DataFrame:
    """
    Parse all history xlsx files and return BGLT bond rate data:
    columns = date, nemotecnico, fec_vcto, years_to_maturity,
              cv_tasa_cierre, cv_tasa_min, cv_tasa_max, cv_monto, cv_num_opes
    Only rows where the bond actually traded (cv_tasa_cierre not null/zero).
    """
    records = []
    for date_str, path in sorted(history_files.items()):
        try:
            wb      = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ms_rows = list(wb["RF-Mercado Secundario"].iter_rows(values_only=True))
            wb.close()
            ms_data = rf_mercado_secundario.parse(ms_rows)
        except Exception:
            continue

        df = ms_data["pub_registro"].copy()
        mask = (
            df["descripcion"].str.lower().str.contains("ext dolares", na=False)
            & df["cv_tasa_cierre"].notna()
            & (df["cv_tasa_cierre"] != 0)
        )
        df = df[mask].copy()
        if df.empty:
            continue

        df["date"] = date_str
        records.append(df[[
            "date", "nemotecnico", "fec_vcto",
            "cv_tasa_cierre", "cv_tasa_min", "cv_tasa_max",
            "cv_monto", "cv_num_opes",
        ]])

    if not records:
        return pd.DataFrame(columns=[
            "date", "nemotecnico", "fec_vcto",
            "cv_tasa_cierre", "cv_tasa_min", "cv_tasa_max",
            "cv_monto", "cv_num_opes", "years_to_maturity",
        ])

    out = pd.concat(records, ignore_index=True)
    out["date"]    = pd.to_datetime(out["date"])
    out["fec_vcto"] = pd.to_datetime(out["fec_vcto"], errors="coerce")
    out["years_to_maturity"] = (out["fec_vcto"] - out["date"]).dt.days / 365.25
    return out


@st.cache_data(show_spinner=False)
def _build_history_df(_files_key: tuple) -> pd.DataFrame:
    """
    Parse all history xlsx files and return a long-format DataFrame:
    columns = date, sector, comprado, vendido, net
    Cache key includes file paths so new files bust the cache automatically.
    """
    records = []
    for date_str, path in sorted(history_files.items()):
        try:
            wb       = openpyxl.load_workbook(path, read_only=True, data_only=True)
            pti_rows = list(wb["RF-PorTipoInver"].iter_rows(values_only=True))
            wb.close()
            pti = rf_por_tipo_inver.parse(pti_rows)
        except Exception:
            continue

        def _ext_col(df: pd.DataFrame) -> "str | None":
            if _EXT_DEBT_INSTR in df.columns:
                return _EXT_DEBT_INSTR
            for c in df.columns:
                if "external" in c.lower() and "dollar" in c.lower():
                    return c
            return None

        bc = _ext_col(pti["buyers"])
        sc = _ext_col(pti["sellers"])
        if bc is None or sc is None:
            continue

        b = pti["buyers"][["sector", bc]].rename(columns={bc: "comprado"})
        s = pti["sellers"][["sector", sc]].rename(columns={sc: "vendido"})
        m = pd.merge(b, s, on="sector", how="outer").fillna(0)
        m["net"]  = m["comprado"] - m["vendido"]
        m["date"] = date_str
        records.append(m)

    if not records:
        return pd.DataFrame(columns=["date", "sector", "comprado", "vendido", "net"])
    return pd.concat(records, ignore_index=True)



# ── Build date catalogue: history files + live API ────────────────────────────
# History files saved by GitHub Actions (data/history/*.xlsx)
def _date_from_xlsx(p: Path) -> str:
    parts = p.stem.split("_")
    try:
        return f"{parts[1]}-{parts[2]}-{parts[3]}"
    except IndexError:
        return p.stem

history_files = {
    _date_from_xlsx(p): str(p)
    for p in sorted(_HISTORY_DIR.glob("BoletinDiario_*.xlsx"))
    if len(p.stem.split("_")) >= 4
}

# Live API dates (re-checked every hour)
with st.spinner("Consultando BVC..."):
    try:
        reports    = _get_reports()
        api_map    = {r["date"]: r["attached"]["url"] for r in reports}
    except Exception as e:
        st.warning(f"No se pudo conectar con la API de BVC: {e}")
        api_map = {}

# Merge: history takes priority; API fills in recent dates not yet saved
all_dates   = sorted(set(history_files) | set(api_map), reverse=True)
today       = date.today().isoformat()

if not all_dates:
    st.error("No hay datos disponibles. Revisa la conexión o el repositorio.")
    st.stop()

# ── Sidebar ────────────────────────────────────────────────────────────────────
chosen_date = st.sidebar.selectbox("📅 Fecha del informe", all_dates)

if today not in all_dates:
    st.sidebar.info(
        f"ℹ️ El boletín de hoy ({today}) aún no está disponible.  \n"
        "Mostrando el más reciente."
    )

st.sidebar.markdown("---")
st.sidebar.markdown(f"**{len(all_dates)} fechas disponibles**")
st.sidebar.markdown("BVC · RF-Mercado Secundario  \nRF-PorTipoInver")

# ── TRM ────────────────────────────────────────────────────────────────────────
trm_data = _get_trm()
trm      = trm_data["valor"]
st.sidebar.markdown("---")
st.sidebar.markdown(
    f"**TRM** `{trm:,.2f}` COP/USD  \n"
    f"<small>Vigencia: {trm_data['vigencia']}  \n"
    f"Fuente: Superfinanciera · datos.gov.co</small>",
    unsafe_allow_html=True,
)

# ── Load & parse ───────────────────────────────────────────────────────────────
with st.spinner("Cargando datos..."):
    try:
        if chosen_date in history_files:
            bond_df, buyers_df, sellers_df, grand_total, report_bytes = _load_from_file(
                history_files[chosen_date], chosen_date
            )
        else:
            bond_df, buyers_df, sellers_df, grand_total, report_bytes = _load(
                api_map[chosen_date], chosen_date
            )
    except Exception as e:
        st.error(f"Error procesando el informe: {e}")
        st.stop()

# ── Header ─────────────────────────────────────────────────────────────────────
st.title("📊 Bonos Deuda Pública Externa USD")
st.caption(f"Informe diario · {chosen_date} · Sistema de Registro (OTC)")
st.info(
    "Este dashboard muestra las negociaciones OTC de Bonos de Deuda Pública Externa "
    "denominados en dólares (BGLT) registradas en la BVC. "
    "**Vista Diaria** detalla la sesión seleccionada; "
    "**Posiciones Históricas** analiza tendencias y patrones entre sesiones.",
    icon="ℹ️",
)

if grand_total == 0:
    st.warning(
        f"No hubo negociaciones de Bonos Deuda Pública Externa USD el **{chosen_date}**. "
        "Selecciona otra fecha en el panel izquierdo."
    )
    st.stop()

col_metric, col_download = st.columns([3, 1])
with col_metric:
    total_usd = grand_total / trm
    st.metric(label="Monto Total Negociado (COP)", value=f"${grand_total:,.0f} M")
    st.caption(f"≈ USD {total_usd:,.2f} M · TRM {trm:,.2f}")
with col_download:
    st.download_button(
        label="⬇️ Descargar Informe Excel",
        data=report_bytes,
        file_name=f"RF_Report_{chosen_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── Tabs ─────────────────────────────────────────────────────────────────────
tab_daily, tab_hist, tab_yield = st.tabs(
    ["📅 Vista Diaria", "📈 Posiciones Históricas", "📉 Curva de Tasas"]
)

# ─────────────────────────────────────────────────────────────────────────────
with tab_daily:

    # ── Bond detail ──────────────────────────────────────────────────────────
    st.subheader("Bonos BGLT")
    st.caption(
        "Detalle por nemotécnico de las operaciones de Compra-Venta registradas en el día. "
        "Montos en millones de COP · Equivalente USD calculado con la TRM del día · Tasas en % efectivo anual."
    )

    # Add USD equivalent column using current TRM
    bond_display = bond_df.copy()
    cop_numeric  = bond_display["CV Monto"].str.replace(",", "", regex=False).apply(pd.to_numeric, errors="coerce")
    usd_col_pos  = bond_display.columns.get_loc("CV Monto") + 1
    bond_display.insert(usd_col_pos, "Monto USD (M)",
        (cop_numeric / trm).apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "")
    )

    st.dataframe(
        bond_display,
        use_container_width=True,
        hide_index=True,
        column_config={
            "CV Monto":      st.column_config.TextColumn("CV Monto (M COP)"),
            "Monto USD (M)": st.column_config.TextColumn("Monto USD (M)"),
            "# Oper.":       st.column_config.NumberColumn("# Oper.",     format="%d"),
            "Tasa Mín":      st.column_config.NumberColumn("Tasa Mín",    format="%.4f"),
            "Tasa Máx":      st.column_config.NumberColumn("Tasa Máx",    format="%.4f"),
            "Tasa Cierre":   st.column_config.NumberColumn("Tasa Cierre", format="%.4f"),
            "Vencimiento":   st.column_config.TextColumn("Vencimiento"),
        },
    )

    st.markdown("---")

    # ── Buyers / Sellers ─────────────────────────────────────────────────────
    st.subheader("Distribución por Tipo de Inversionista")
    st.caption(
        "Muestra qué sectores compraron y vendieron BGLT en la sesión. "
        "Un sector con alto monto comprado y bajo vendido es **demandante neto** de papel; "
        "el caso inverso indica **distribución**. "
        "Los sectores sin operaciones no participaron en BGLT ese día."
    )
    col_b, col_s = st.columns(2)

    def _render_side(col, df: pd.DataFrame, monto_col: str, title: str):
        with col:
            st.markdown(f"#### {title}")
            active = df[df[monto_col] > 0]
            zero   = df[df[monto_col] == 0]
            display = active[[monto_col, "Sector"]].copy()
            display[monto_col] = display[monto_col].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
            st.dataframe(
                display,
                use_container_width=True,
                hide_index=True,
                column_config={monto_col: st.column_config.TextColumn(monto_col)},
            )
            st.caption(
                f"**Total: ${active[monto_col].sum():,.0f} M** · "
                f"{len(active)} sector(es) activo(s) · "
                f"{len(zero)} sin operaciones"
            )
            if not zero.empty:
                with st.expander(f"Sectores sin operaciones ({len(zero)})"):
                    st.dataframe(zero[["Sector"]], use_container_width=True, hide_index=True)

    _render_side(col_b, buyers_df,  "Monto Comprado", "🟢 Compradores")
    _render_side(col_s, sellers_df, "Monto Vendido",  "🔴 Vendedores")


# ─────────────────────────────────────────────────────────────────────────────
with tab_hist:
    st.subheader("Posiciones Netas por Sector — Bonos Ext. USD")
    st.caption(
        "Positivo = comprador neto · Negativo = vendedor neto · "
        "Solo sectores con actividad en al menos una sesión"
    )

    with st.spinner("Construyendo serie histórica..."):
        hist_df = _build_history_df(tuple(sorted(history_files.items())))

    if hist_df.empty:
        st.info("No hay datos históricos disponibles aún.")
    else:
        # ── Prep ─────────────────────────────────────────────────────────
        active_sectors = (
            hist_df.groupby("sector")["net"]
            .apply(lambda x: (x.abs() > 0).any())
            .loc[lambda x: x]
            .index.tolist()
        )
        h = hist_df[hist_df["sector"].isin(active_sectors)].copy()
        h["date"]     = pd.to_datetime(h["date"])
        h["sector_s"] = h["sector"].str.split("/").str[0].str.strip()

        # ── KPIs ──────────────────────────────────────────────────────────
        cum        = h.groupby("sector_s")["net"].sum().sort_values(ascending=False)
        n_days     = int(h["date"].nunique())
        top_buyer  = cum.index[0]  if len(cum) > 0 else "—"
        top_seller = cum.index[-1] if len(cum) > 0 else "—"

        k1, k2, k3 = st.columns(3)
        k1.metric("📅 Fechas analizadas",    n_days)
        k2.metric("🟢 Mayor comprador neto", top_buyer)
        k3.metric("🔴 Mayor vendedor neto",  top_seller)

        st.markdown("---")

        # ── Line chart: net position per day per sector ───────────────────
        st.markdown("##### Posición Neta Diaria por Sector")
        st.caption(
            "Cada línea es **comprado − vendido** por sector en cada sesión. "
            "Valores **sobre cero** → comprador neto ese día; "
            "**bajo cero** → vendedor neto. "
            "Hover sobre un punto para ver el detalle de compras y ventas."
        )

        zero_rule = (
            alt.Chart(pd.DataFrame({"y": [0]}))
            .mark_rule(color="#666", strokeDash=[4, 4])
            .encode(y="y:Q")
        )
        line_chart = (
            alt.Chart(h)
            .mark_line(point=True, strokeWidth=2)
            .encode(
                x=alt.X("date:T", title="Fecha",
                        axis=alt.Axis(format="%d %b", labelAngle=-30)),
                y=alt.Y("net:Q",  title="Posición Neta (M COP)",
                        axis=alt.Axis(format=",.0f")),
                color=alt.Color("sector_s:N", title="Sector"),
                tooltip=[
                    alt.Tooltip("date:T",     title="Fecha",         format="%Y-%m-%d"),
                    alt.Tooltip("sector_s:N", title="Sector"),
                    alt.Tooltip("net:Q",      title="Pos. Neta",     format=",.0f"),
                    alt.Tooltip("comprado:Q", title="Comprado",      format=",.0f"),
                    alt.Tooltip("vendido:Q",  title="Vendido",       format=",.0f"),
                ],
            )
            .properties(height=380)
            .interactive()
        )
        st.altair_chart(zero_rule + line_chart, use_container_width=True)

        st.markdown("---")

        # ── Heatmap: sector × date ────────────────────────────────────────
        st.markdown("##### Mapa de Calor — Posición Neta (M COP)")
        st.caption(
            "Tabla cruzada sector × fecha. "
            "🟢 Verde intenso = gran comprador neto · 🔴 Rojo intenso = gran vendedor neto · Blanco/amarillo = neutral. "
            "Útil para detectar **patrones recurrentes**: p.ej. si un sector aparece siempre en rojo, "
            "es un distribuidor estructural de BGLT."
        )

        pivot = h.pivot_table(
            index="sector_s", columns="date", values="net",
            aggfunc="sum", fill_value=0,
        )
        pivot.columns = [d.strftime("%d/%m") for d in pivot.columns]
        pivot.index.name = "Sector"
        abs_max = float(pivot.abs().max().max()) or 1.0
        st.dataframe(
            pivot.style
            .background_gradient(cmap="RdYlGn", axis=None, vmin=-abs_max, vmax=abs_max)
            .format("{:,.0f}"),
            use_container_width=True,
        )

        st.markdown("---")

        # ── Cumulative net bar ────────────────────────────────────────────
        st.markdown("##### Posición Neta Acumulada en el Período")
        st.caption(
            "Suma de posiciones netas diarias en todas las sesiones analizadas. "
            "Identifica los **compradores y vendedores estructurales** de BGLT: "
            "barras verdes = acumulación neta de papel; barras rojas = distribución neta."
        )

        cum_df = (
            h.groupby("sector_s")["net"]
            .sum()
            .reset_index()
            .rename(columns={"sector_s": "Sector", "net": "Neto"})
            .sort_values("Neto", ascending=False)
        )
        cum_df["Rol"] = cum_df["Neto"].apply(
            lambda x: "Comprador" if x >= 0 else "Vendedor"
        )

        bar = (
            alt.Chart(cum_df)
            .mark_bar(cornerRadiusEnd=3)
            .encode(
                x=alt.X("Neto:Q", title="Posición Neta Acumulada (M COP)",
                        axis=alt.Axis(format=",.0f")),
                y=alt.Y("Sector:N", sort="-x", title=None),
                color=alt.Color(
                    "Rol:N",
                    scale=alt.Scale(
                        domain=["Comprador", "Vendedor"],
                        range=["#2ecc71", "#e74c3c"],
                    ),
                    legend=None,
                ),
                tooltip=[
                    alt.Tooltip("Sector:N"),
                    alt.Tooltip("Neto:Q", title="Posición Neta", format=",.0f"),
                    alt.Tooltip("Rol:N"),
                ],
            )
            .properties(height=max(200, len(cum_df) * 45))
        )
        st.altair_chart(bar, use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
with tab_yield:
    st.subheader("Curva de Tasas — Bonos BGLT")
    st.caption(
        "Análisis de tasas de cierre (% e.a.) de los bonos de Deuda Pública Externa USD. "
        "Usa la fecha del panel izquierdo para ver el snapshot de curva de ese día."
    )

    with st.spinner("Cargando datos de tasas..."):
        yield_df = _build_yield_history_df(tuple(sorted(history_files.items())))

    if yield_df.empty:
        st.info("No hay datos de tasas disponibles aún.")
    else:
        # ── Chart 1: Yield Curve snapshot ──────────────────────────────────
        snap = yield_df[yield_df["date"] == pd.to_datetime(chosen_date)]
        if snap.empty:
            # Fall back to the most recent available session
            snap      = yield_df[yield_df["date"] == yield_df["date"].max()]
            snap_date = yield_df["date"].max().strftime("%Y-%m-%d")
            st.caption(f"⚠️ Sin datos para {chosen_date}. Mostrando la sesión más reciente: {snap_date}.")
        else:
            snap_date = chosen_date

        st.markdown(f"##### 📈 Curva de Tasas al {snap_date}")
        st.caption(
            "Cada punto es un bono BGLT que negóció en esa sesión. "
            "**Eje X** = años al vencimiento · **Eje Y** = tasa de cierre (% e.a.). "
            "El tamaño del punto refleja el monto negociado. "
            "La línea punteada conecta los puntos ordenados por plazo, aproximando la forma de la curva."
        )

        # Dashed line connecting dots ordered by maturity
        curve_line = (
            alt.Chart(snap.sort_values("years_to_maturity"))
            .mark_line(strokeDash=[5, 3], color="#888", strokeWidth=1.5)
            .encode(
                x=alt.X("years_to_maturity:Q"),
                y=alt.Y("cv_tasa_cierre:Q"),
            )
        )
        # Dots sized by monto, colored by bond
        dots = (
            alt.Chart(snap)
            .mark_circle(opacity=0.9)
            .encode(
                x=alt.X("years_to_maturity:Q", title="Años al Vencimiento",
                        scale=alt.Scale(zero=False)),
                y=alt.Y("cv_tasa_cierre:Q", title="Tasa de Cierre (% e.a.)",
                        scale=alt.Scale(zero=False),
                        axis=alt.Axis(format=".4f")),
                size=alt.Size("cv_monto:Q", title="Monto (M COP)",
                              scale=alt.Scale(range=[150, 900])),
                color=alt.Color("nemotecnico:N", title="Bono"),
                tooltip=[
                    alt.Tooltip("nemotecnico:N",       title="Bono"),
                    alt.Tooltip("years_to_maturity:Q", title="Años al Vto.", format=".1f"),
                    alt.Tooltip("cv_tasa_cierre:Q",    title="Tasa Cierre",  format=".4f"),
                    alt.Tooltip("cv_tasa_min:Q",       title="Tasa Mín",     format=".4f"),
                    alt.Tooltip("cv_tasa_max:Q",       title="Tasa Máx",     format=".4f"),
                    alt.Tooltip("cv_monto:Q",          title="Monto (M COP)",format=",.0f"),
                    alt.Tooltip("cv_num_opes:Q",       title="# Oper."),
                ],
            )
            .properties(height=360)
            .interactive()
        )
        st.altair_chart(curve_line + dots, use_container_width=True)

        st.markdown("---")

        # ── Chart 2: Rate evolution per bond over time ───────────────────
        st.markdown("##### 🗓️ Evolución de Tasas de Cierre por Bono")
        st.caption(
            "Cada línea es un bono BGLT. Solo se grafican días en que el bono negóció. "
            "**Tasa al alza** = el bono se abarata (precio cae, rentabilidad sube). "
            "**Tasa a la baja** = el bono se encarece (precio sube, rentabilidad cae)."
        )

        rate_evo = (
            alt.Chart(yield_df)
            .mark_line(point=True, strokeWidth=2)
            .encode(
                x=alt.X("date:T", title="Fecha",
                        axis=alt.Axis(format="%d %b", labelAngle=-30)),
                y=alt.Y("cv_tasa_cierre:Q", title="Tasa de Cierre (% e.a.)",
                        scale=alt.Scale(zero=False),
                        axis=alt.Axis(format=".4f")),
                color=alt.Color("nemotecnico:N", title="Bono"),
                tooltip=[
                    alt.Tooltip("date:T",            title="Fecha",         format="%Y-%m-%d"),
                    alt.Tooltip("nemotecnico:N",      title="Bono"),
                    alt.Tooltip("cv_tasa_cierre:Q",   title="Tasa Cierre",   format=".4f"),
                    alt.Tooltip("cv_tasa_min:Q",      title="Tasa Mín",      format=".4f"),
                    alt.Tooltip("cv_tasa_max:Q",      title="Tasa Máx",      format=".4f"),
                    alt.Tooltip("cv_monto:Q",         title="Monto (M COP)", format=",.0f"),
                ],
            )
            .properties(height=360)
            .interactive()
        )
        st.altair_chart(rate_evo, use_container_width=True)

        st.markdown("---")

        # ── Table: latest rate per bond ────────────────────────────────────
        st.markdown("##### 📌 Últimas Tasas Registradas por Bono")
        st.caption(
            "La última tasa de cierre observada para cada BGLT, "
            "ordenada de menor a mayor plazo. "
            "Sirve como referencia rápida del nivel de tasas vigente."
        )

        latest = (
            yield_df.sort_values("date")
            .groupby("nemotecnico", as_index=False)
            .last()
            .sort_values("years_to_maturity")
            .rename(columns={
                "nemotecnico":       "Bono",
                "date":              "Última Sesión",
                "fec_vcto":          "Vencimiento",
                "years_to_maturity": "Años al Vto.",
                "cv_tasa_cierre":    "Tasa Cierre",
                "cv_tasa_min":       "Tasa Mín",
                "cv_tasa_max":       "Tasa Máx",
                "cv_monto":          "Monto (M COP)",
                "cv_num_opes":       "# Oper.",
            })
        )
        latest["Última Sesión"] = latest["Última Sesión"].dt.strftime("%Y-%m-%d")
        latest["Vencimiento"]   = latest["Vencimiento"].dt.strftime("%Y-%m-%d")
        latest["Monto (M COP)"] = latest["Monto (M COP)"].apply(
            lambda x: f"{x:,.0f}" if pd.notna(x) else ""
        )

        st.dataframe(
            latest[[
                "Bono", "Última Sesión", "Vencimiento", "Años al Vto.",
                "Tasa Cierre", "Tasa Mín", "Tasa Máx", "Monto (M COP)", "# Oper.",
            ]],
            use_container_width=True,
            hide_index=True,
            column_config={
                "Años al Vto.": st.column_config.NumberColumn(format="%.1f"),
                "Tasa Cierre":  st.column_config.NumberColumn(format="%.4f"),
                "Tasa Mín":     st.column_config.NumberColumn(format="%.4f"),
                "Tasa Máx":     st.column_config.NumberColumn(format="%.4f"),
                "Monto (M COP)": st.column_config.TextColumn("Monto (M COP)"),
            },
        )
