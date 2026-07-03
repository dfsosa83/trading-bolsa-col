"""
RF Daily Report Generator
=========================
Reads the latest (or a specific date's) BoletinDiario xlsx, parses the two
fixed-income sheets, and writes a summary Excel report to data/reports/.

Usage
-----
    # Latest available xlsx:
    python reports/generate_report.py

    # Specific date:
    python reports/generate_report.py --date 2026-06-25

Output
------
    data/reports/RF_Report_YYYY-MM-DD.xlsx
      Sheet 1 – "Merc Sec Transaccional"  : top bonds by traded volume
      Sheet 2 – "Merc Sec Registro"       : top bonds by traded volume
      Sheet 3 – "BGLT - Bonos Ext USD"    : external debt bonds (BGLT) detail + monto sum
      Sheet 4 – "Por Tipo Inver Compras"  : buyers by sector × instrument
      Sheet 5 – "Por Tipo Inver Ventas"   : sellers by sector × instrument
"""

import argparse
import io
import os
import sys
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Make sure the project root is on the path when called directly
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from reports.parsers import rf_mercado_secundario, rf_por_tipo_inver

# ── Paths ─────────────────────────────────────────────────────────────────────
_ROOT      = Path(__file__).resolve().parent.parent
_XLSX_DIR  = _ROOT / "data" / "daily_data" / "xlsx"
_OUT_DIR   = _ROOT / "data" / "reports"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _header_style(ws, row: int, bg_hex: str = "1F4E79") -> None:
    """Bold white text on dark background for a full row."""
    fill = PatternFill("solid", fgColor=bg_hex)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws) -> None:
    """Approximate column widths based on max content length."""
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


def _write_df(ws, df: pd.DataFrame, title: str, start_row: int = 1) -> int:
    """
    Write a DataFrame to a worksheet starting at `start_row`.
    Returns the next available row after the block.
    """
    # Title
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    header_row = start_row + 1

    # Headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name)
    _header_style(ws, header_row)

    # Data
    for row_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, float):
                cell.number_format = "#,##0.00"

    return header_row + len(df) + 2   # +2 for spacing


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_mercado_sec_sheet(wb, name: str, df: pd.DataFrame, report_date: str) -> None:
    ws = wb.create_sheet(name)

    # ── Summary block: top 10 by total traded monto (cv + sim) ──────────────
    df = df.copy()
    df["total_monto"] = df["cv_monto"].fillna(0) + df["sim_monto"].fillna(0)
    top = (
        df[df["total_monto"] > 0]
        .sort_values("total_monto", ascending=False)
        .head(10)
    )

    summary_cols = [
        "nemotecnico", "descripcion", "fec_vcto",
        "cv_monto", "cv_num_opes", "cv_tasa_cierre",
        "sim_monto", "sim_num_opes", "sim_tasa_cierre",
        "total_monto",
    ]
    summary = top[summary_cols].rename(columns={
        "nemotecnico":    "Nemotécnico",
        "descripcion":    "Descripción",
        "fec_vcto":       "Vto.",
        "cv_monto":       "CV Monto",
        "cv_num_opes":    "CV # Opes",
        "cv_tasa_cierre": "CV Tasa Cierre",
        "sim_monto":      "Sim Monto",
        "sim_num_opes":   "Sim # Opes",
        "sim_tasa_cierre":"Sim Tasa Cierre",
        "total_monto":    "Monto Total",
    })

    next_row = _write_df(
        ws, summary,
        title=f"Top 10 por Monto Negociado — {name} ({report_date})",
        start_row=1,
    )

    # ── Totals block ──────────────────────────────────────────────────────────
    totals = pd.DataFrame([{
        "Concepto":     "TOTAL",
        "CV Monto":     df["cv_monto"].fillna(0).sum(),
        "CV # Opes":    int(df["cv_num_opes"].fillna(0).sum()),
        "Sim Monto":    df["sim_monto"].fillna(0).sum(),
        "Sim # Opes":   int(df["sim_num_opes"].fillna(0).sum()),
        "Monto Total":  df["total_monto"].sum(),
    }])
    _write_df(ws, totals, title="Totales Generales", start_row=next_row)
    _autofit(ws)


_EXT_DEBT_INSTR = "PUBLIC EXTERNAL DEBT BONDS DOLLAR DENOMINATED"


def _build_ext_bonds_usd_sheet(
    wb,
    df_registro: pd.DataFrame,
    pti_data: dict,
    report_date: str,
) -> None:
    """
    Single sheet combining three related views for external-debt-in-USD bonds:
      1. Bond detail (BGLT) with traded monto and rates
      2. Buyer breakdown by investor sector (RF-PorTipoInver)
      3. Seller breakdown by investor sector (RF-PorTipoInver)
    All three totals equal the same figure (201,222.08 for 2026-06-25).
    """
    ws = wb.create_sheet("BGLT - Bonos Ext USD")

    # ── Section 1: Bond detail ─────────────────────────────────────────────
    mask = df_registro["descripcion"].str.lower().str.contains("ext dolares", na=False)
    df_bonds = df_registro[mask].copy()
    df_bonds["total_monto"] = df_bonds["cv_monto"].fillna(0) + df_bonds["sim_monto"].fillna(0)

    # Monto Total first, then identifier columns, then rates
    bond_cols = {
        "total_monto":    "Monto Total",
        "nemotecnico":    "Nemotécnico",
        "descripcion":    "Descripción",
        "fec_emision":    "Fec. Emisión",
        "fec_vcto":       "Vto.",
        "cv_monto":       "CV Monto",
        "cv_num_opes":    "CV # Opes",
        "cv_tasa_min":    "Tasa Mín",
        "cv_tasa_max":    "Tasa Máx",
        "cv_tasa_med":    "Tasa Med",
        "cv_tasa_cierre": "Tasa Cierre",
        "sim_monto":      "Sim Monto",
        "sim_num_opes":   "Sim # Opes",
    }
    detail = df_bonds[list(bond_cols)].rename(columns=bond_cols)
    next_row = _write_df(
        ws, detail,
        title=f"Bonos Deuda Pública Externa USD (BGLT) — {report_date}",
        start_row=1,
    )

    bond_total = df_bonds["total_monto"].sum()
    bond_totals = pd.DataFrame([{
        "Concepto":    "TOTAL",
        "CV Monto":    df_bonds["cv_monto"].fillna(0).sum(),
        "CV # Opes":   int(df_bonds["cv_num_opes"].fillna(0).sum()),
        "Sim Monto":   df_bonds["sim_monto"].fillna(0).sum(),
        "Sim # Opes":  int(df_bonds["sim_num_opes"].fillna(0).sum()),
        "Monto Total": bond_total,
    }])
    next_row = _write_df(ws, bond_totals, title="Totales Bonos", start_row=next_row)

    # ── Section 2: Compradores (Buyers) — all sectors ────────────────────
    buyers_df = pti_data["buyers"]
    if _EXT_DEBT_INSTR in buyers_df.columns:
        buyers = (
            buyers_df[["sector", _EXT_DEBT_INSTR]]
            .rename(columns={"sector": "Sector", _EXT_DEBT_INSTR: "Monto Comprado"})
            .sort_values("Monto Comprado", ascending=False)
            .reset_index(drop=True)
        )
        buyers = buyers[["Monto Comprado", "Sector"]]
    else:
        buyers = pd.DataFrame([{"Monto Comprado": 0.0, "Sector": "(sin operaciones)"}])
    buyers_total = pd.DataFrame([{"Monto Comprado": buyers["Monto Comprado"].sum(), "Sector": "TOTAL"}])
    buyers_display = pd.concat([buyers, buyers_total], ignore_index=True)
    next_row = _write_df(
        ws, buyers_display,
        title=f"Compradores por Sector — Bonos Ext. USD  (total: {bond_total:,.2f})",
        start_row=next_row,
    )

    # ── Section 3: Vendedores (Sellers) — all sectors ────────────────────
    sellers_df = pti_data["sellers"]
    if _EXT_DEBT_INSTR in sellers_df.columns:
        sellers = (
            sellers_df[["sector", _EXT_DEBT_INSTR]]
            .rename(columns={"sector": "Sector", _EXT_DEBT_INSTR: "Monto Vendido"})
            .sort_values("Monto Vendido", ascending=False)
            .reset_index(drop=True)
        )
        sellers = sellers[["Monto Vendido", "Sector"]]
    else:
        sellers = pd.DataFrame([{"Monto Vendido": 0.0, "Sector": "(sin operaciones)"}])
    sellers_total = pd.DataFrame([{"Monto Vendido": sellers["Monto Vendido"].sum(), "Sector": "TOTAL"}])
    sellers_display = pd.concat([sellers, sellers_total], ignore_index=True)
    _write_df(
        ws, sellers_display,
        title=f"Vendedores por Sector — Bonos Ext. USD  (total: {bond_total:,.2f})",
        start_row=next_row,
    )

    _autofit(ws)


def _build_tipo_inver_sheet(wb, name: str, df: pd.DataFrame, label: str, report_date: str) -> None:
    ws = wb.create_sheet(name)
    df_disp = df.rename(columns={"sector": "Sector", "total": "Total"})
    _write_df(
        ws, df_disp,
        title=f"{label} por Sector e Instrumento ({report_date})",
        start_row=1,
    )
    _autofit(ws)


# ── Main orchestration ────────────────────────────────────────────────────────

def _latest_xlsx() -> Path:
    files = sorted(_XLSX_DIR.glob("BoletinDiario_*.xlsx"), reverse=True)
    if not files:
        raise FileNotFoundError(f"No xlsx files found in {_XLSX_DIR}")
    return files[0]


def _date_from_path(path: Path) -> str:
    """Extract YYYY-MM-DD from filename like BoletinDiario_2026_06_25.xlsx."""
    stem = path.stem  # e.g. "BoletinDiario_2026_06_25"
    parts = stem.split("_")
    # parts: ['BoletinDiario', 'YYYY', 'MM', 'DD', ...]
    try:
        return f"{parts[1]}-{parts[2]}-{parts[3]}"
    except IndexError:
        return stem


def _load_ws_rows(xlsx_path: Path, sheet_name: str) -> list[tuple]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def generate(xlsx_path: Path) -> Path:
    report_date = _date_from_path(xlsx_path)
    print(f"  Source : {xlsx_path.name}")
    print(f"  Date   : {report_date}")

    # ── Parse ─────────────────────────────────────────────────────────────────
    print("  Parsing RF-Mercado Secundario ...")
    ms_rows  = _load_ws_rows(xlsx_path, "RF-Mercado Secundario")
    ms_data  = rf_mercado_secundario.parse(ms_rows)

    print("  Parsing RF-PorTipoInver ...")
    pti_rows = _load_ws_rows(xlsx_path, "RF-PorTipoInver")
    pti_data = rf_por_tipo_inver.parse(pti_rows)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    _build_mercado_sec_sheet(wb, "Merc Sec Transaccional", ms_data["pub_transaccional"], report_date)
    _build_mercado_sec_sheet(wb, "Merc Sec Registro",      ms_data["pub_registro"],      report_date)
    _build_ext_bonds_usd_sheet(wb, ms_data["pub_registro"], pti_data, report_date)
    _build_tipo_inver_sheet( wb, "Por Tipo Inver Compras", pti_data["buyers"],  "Compradores", report_date)
    _build_tipo_inver_sheet( wb, "Por Tipo Inver Ventas",  pti_data["sellers"], "Vendedores",  report_date)

    # ── Save ──────────────────────────────────────────────────────────────────
    _OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = _OUT_DIR / f"RF_Report_{report_date}.xlsx"
    wb.save(out_path)
    print(f"  Report : {out_path}")
    return out_path


def generate_to_bytes(xlsx_bytes: bytes, report_date: str) -> io.BytesIO:
    """Same as generate() but works entirely in memory — no files read or written.

    Parameters
    ----------
    xlsx_bytes   : raw bytes of the BoletinDiario xlsx file
    report_date  : 'YYYY-MM-DD' string used for sheet titles

    Returns
    -------
    io.BytesIO positioned at 0, ready for st.download_button or further use.
    """
    wb_in = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ms_rows  = list(wb_in["RF-Mercado Secundario"].iter_rows(values_only=True))
    pti_rows = list(wb_in["RF-PorTipoInver"].iter_rows(values_only=True))
    wb_in.close()

    ms_data  = rf_mercado_secundario.parse(ms_rows)
    pti_data = rf_por_tipo_inver.parse(pti_rows)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    _build_mercado_sec_sheet(wb_out, "Merc Sec Transaccional", ms_data["pub_transaccional"], report_date)
    _build_mercado_sec_sheet(wb_out, "Merc Sec Registro",      ms_data["pub_registro"],      report_date)
    _build_ext_bonds_usd_sheet(wb_out, ms_data["pub_registro"], pti_data, report_date)
    _build_tipo_inver_sheet(wb_out, "Por Tipo Inver Compras",   pti_data["buyers"],  "Compradores", report_date)
    _build_tipo_inver_sheet(wb_out, "Por Tipo Inver Ventas",    pti_data["sellers"], "Vendedores",  report_date)

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate RF daily report.")
    parser.add_argument("--date", default=None, help="YYYY-MM-DD (default: latest)")
    args = parser.parse_args()

    if args.date:
        date_str = args.date.replace("-", "_")
        matches = list(_XLSX_DIR.glob(f"BoletinDiario_{date_str}*.xlsx"))
        if not matches:
            print(f"No xlsx found for date {args.date} in {_XLSX_DIR}")
            sys.exit(1)
        xlsx_path = matches[0]
    else:
        xlsx_path = _latest_xlsx()

    print("Generating RF report...")
    out = generate(xlsx_path)
    print("Done.")


if __name__ == "__main__":
    main()
